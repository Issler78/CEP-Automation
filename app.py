from time import sleep
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

#abrir a tabela
endereços_book = openpyxl.load_workbook('enderecos.xlsx')
endereços_page = endereços_book['Sheet1']

#nova tabela
enderecos_completos = openpyxl.Workbook()
enderecos_completos_page = enderecos_completos['Sheet']
enderecos_completos_page.append([
    'Nome',
    'Endereço',
    'CEP',
    'Cidade',
    'Estado'
])



#entrar no site para buscar o cep
options = webdriver.EdgeOptions()
options.add_experimental_option('detach', True)
driver = webdriver.Edge(options=options)

#entrar no link do site
driver.get('https://www2.correios.com.br/sistemas/buscacep/BuscaCepEndereco.cfm')
sleep(2)

#percorrer toda a tabela de enderecos
for row in endereços_page.iter_rows(values_only=True, min_row=2):
    nome, endereco, cep = row
    
    #pesquisar o input de cep no site
    sleep(0.5)
    cep_input = driver.find_element(By.XPATH, "//input[@name='relaxation']")
    cep_input.send_keys(cep)
    
    sleep(0.5)
    #pesquisar botao de busca
    busca_input = driver.find_element(By.XPATH, "//input[@class='btn2 float-right']")
    busca_input.click()
    sleep(0.5)
    
    #pesquisar a tabela onde esta os dados do resultado
    tds_tabela = driver.find_elements(By.XPATH, "//table[@class='tmptabela']//td")
    for ind, td in enumerate(tds_tabela):
        #se o dado for a cidade/UF
        if ind == 2:
            cidade_estado = td.text.split('/')
            cidade, estado = cidade_estado
    
    #salvar na tabela
    enderecos_completos_page.append([
        nome,
        endereco,
        cep,
        cidade,
        estado
    ])
    
    #achar o link para retornar para a pesquisa
    sleep(0.5)
    link_retornar = driver.find_element(By.XPATH, "//a[@href='sistemas/buscacep/BuscaCepEndereco.cfm']")
    link_retornar.click()
#salvar tabela
enderecos_completos.save('enderecos_completos.xlsx')
driver.quit()